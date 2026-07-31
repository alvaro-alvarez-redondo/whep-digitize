#!/usr/bin/env python
"""Full-dataset R-vs-Python parity harness.

Automates the previously-manual end-to-end parity procedure: run the R pipeline
(``whep-digitalization``) and the Python pipeline (:func:`whep_digitize.pipeline.run_pipeline`)
over the **same full production dataset**, then diff their outputs:

* **processed TSVs** — byte-level;
* **unique-list workbooks** — content-level (sheet names + cell values), because R ``writexl``
  and Python ``xlsxwriter`` never produce identical ZIP bytes.

Every difference must be an instance of the single accepted divergence: the **normalization
policy** (NFD diacritic strip) deliberately replacing R's ICU ``Latin-ASCII`` transliteration —
see ``.claude/docs/r-to-python-mapping.md`` risk #1. On the production dataset that is 13 rows,
all of them ICU rendering a trailing ``(R)`` for ``®`` that the policy drops (``philippines r``
-> ``philippines`` and the same for ``nicaragus``/``australia``/``brazil``/``uruguay``), with
**value sums and row counts unchanged**. Anything else — an extra differing row, a changed
``value``/``year``, row-count or sum drift — exits non-zero.

**Rows are compared as a multiset, not positionally.** An accepted divergence rewrites a sort
key (``philippines r`` and ``philippines`` sort to different places), so every later row shifts
and a line-by-line diff degenerates into tens of thousands of meaningless pairings — some of
which coincidentally look like valid divergences. Instead the two row multisets are differenced,
and the leftover rows are matched 1:1 against each other; only a matched pair whose every
differing field has the accepted shape counts as an accepted divergence.

**Precondition (hard).** R is the parity oracle; there is no way to diff against it without it.
This harness requires the sibling R project, an executable R, and the full raw dataset. When any
of those is missing it exits ``2`` with an explanatory message rather than silently passing.

Usage::

    .venv/Scripts/python.exe scripts/parity_full_dataset.py            # reuse R's existing output
    .venv/Scripts/python.exe scripts/parity_full_dataset.py --run-r    # regenerate it first

``--run-r`` executes the R project's documented entry point (``whep-digitalization.R``), which
**overwrites that project's** ``data/2-postpro`` **and** ``data/3-export`` **outputs**; it is
opt-in for exactly that reason.

Environment overrides: ``WHEP_R_PROJECT`` (R project root), ``WHEP_R_HOME`` (R installation).
"""

from __future__ import annotations

import argparse
import os
import shutil
import subprocess
import sys
import tempfile
from collections import Counter
from collections.abc import Iterable, Sequence
from dataclasses import dataclass, field
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

_REPO_ROOT = Path(__file__).resolve().parents[1]
_DEFAULT_R_PROJECT = _REPO_ROOT.parent / "whep-digitalization"
_DEFAULT_R_HOME = Path("C:/Program Files/R/R-4.6.0")

_R_ENTRY_POINT = "whep-digitalization.R"

# R data layout -> this project's import-layer names (mirrors .claude/bench/bench.py).
_R_LAYER_MAP = {
    "10-raw_import": "raw",
    "11-clean_import": "clean",
    "12-standardize_import": "standardize",
    "13-harmonize_import": "harmonize",
}

_R_RAW_LAYER = "10-raw_import"
_R_PROCESSED_DIR = Path("data") / "3-export" / "processed_data"
_R_LISTS_DIR = Path("data") / "3-export" / "lists"
_PY_PROCESSED_DIR = Path("data") / "export" / "processed"
_PY_LISTS_DIR = Path("data") / "export" / "lists"

#: Columns a normalization divergence may never touch — a difference here is always a defect.
_INVARIANT_COLUMNS = frozenset({"year", "value"})

#: Accepted differing-row budget on the production dataset (risk #1).
_DEFAULT_DIVERGENCE_BUDGET = 13

#: Beyond this many unmatched rows, classification is pointless (and quadratic) — fail outright.
_CLASSIFY_LIMIT_FACTOR = 10

Row = tuple[str, ...]


class PreconditionError(RuntimeError):
    """Raised when the R reference cannot be produced or located."""


@dataclass
class RowDivergence:
    """One row that differs between R and Python, with its per-field detail.

    Attributes:
        location: The artifact the row belongs to (``"<file>"`` or ``"<file>!<sheet>"``).
        fields: Rendered ``"<column>: <r> -> <py>"`` descriptions of the differing fields.
    """

    location: str
    fields: list[str]

    def render(self) -> str:
        """Render the divergence as a single line.

        Returns:
            The location followed by its differing fields.
        """
        return f"{self.location}: " + "; ".join(self.fields)


@dataclass
class ComparisonReport:
    """Outcome of comparing one artifact class (processed TSVs or list workbooks).

    Attributes:
        label: The artifact class being compared.
        identical: Files that matched exactly.
        accepted: Rows differing only by the accepted normalization policy.
        rejected: Rows that differ for any other reason.
        errors: Structural failures (missing files, row-count drift, sum drift).
    """

    label: str
    identical: list[str] = field(default_factory=list)
    accepted: list[RowDivergence] = field(default_factory=list)
    rejected: list[RowDivergence] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)


# ---------------------------------------------------------------------------
# Divergence classification
# ---------------------------------------------------------------------------


def is_subsequence(needle: str, haystack: str) -> bool:
    """Report whether ``needle`` can be obtained from ``haystack`` by deleting characters.

    Args:
        needle: The candidate subsequence.
        haystack: The string to delete characters from.

    Returns:
        ``True`` when every character of ``needle`` appears in ``haystack`` in order.
    """
    it = iter(haystack)
    return all(ch in it for ch in needle)


def classify_difference(r_value: str, py_value: str) -> tuple[bool, str]:
    """Decide whether an R-vs-Python cell difference is the accepted normalization divergence.

    The policy (NFD diacritic strip, then ``[^a-z0-9]+`` -> space) **drops** the characters ICU
    ``Latin-ASCII`` **expands or maps** (``(R)`` for ``®``, ``1/2`` for ``½``, ``ss`` for ``ß``,
    ``ae`` for ``æ``, ...). The Python value is therefore always reachable from the R value by
    deleting characters. A genuine defect — a wrong join, a shifted field, a dropped record —
    effectively never satisfies that relation, which is what makes it a usable discriminator.

    Blanking a non-empty value is rejected outright: the subsequence relation holds trivially for
    the empty string, so allowing it would let a whole class of real defects through.

    Args:
        r_value: The value R produced.
        py_value: The value Python produced.

    Returns:
        ``(accepted, reason)`` — ``reason`` explains the rejection and is empty when accepted.
    """
    if r_value == py_value:
        return True, ""
    if py_value == "" and r_value != "":
        return False, "Python blanked a value R populated"
    if not is_subsequence(py_value, r_value):
        return False, "Python value is not reachable from R's by deleting characters"
    return True, ""


def describe_row_difference(
    r_row: Row, py_row: Row, columns: Sequence[str]
) -> tuple[bool, list[str]]:
    """Compare two rows field by field and decide whether the whole row is an accepted divergence.

    Args:
        r_row: The R row.
        py_row: The Python row.
        columns: Column names; positions beyond its length are named ``colN``.

    Returns:
        ``(accepted, descriptions)`` — one description per differing field.
    """
    if len(r_row) != len(py_row):
        return False, [f"field count {len(r_row)} vs {len(py_row)}"]
    accepted = True
    descriptions: list[str] = []
    for index, (r_cell, py_cell) in enumerate(zip(r_row, py_row, strict=True)):
        if r_cell == py_cell:
            continue
        name = columns[index] if index < len(columns) else f"col{index}"
        if name in _INVARIANT_COLUMNS:
            ok, reason = False, f"{name} must never differ"
        else:
            ok, reason = classify_difference(r_cell, py_cell)
        accepted = accepted and ok
        suffix = "" if ok else f"  <-- {reason}"
        descriptions.append(f"{name}: {r_cell!r} -> {py_cell!r}{suffix}")
    return accepted, descriptions


def _maximum_matching(adjacency: Sequence[Sequence[int]], right_size: int) -> list[int]:
    """Find a maximum bipartite matching (Kuhn's augmenting-path algorithm).

    Args:
        adjacency: For each left node, the right nodes it may pair with.
        right_size: Number of right nodes.

    Returns:
        ``match_right[j]`` = the left node matched to right node ``j``, or ``-1``.
    """
    match_right = [-1] * right_size

    def augment(left: int, seen: list[bool]) -> bool:
        for right in adjacency[left]:
            if seen[right]:
                continue
            seen[right] = True
            if match_right[right] == -1 or augment(match_right[right], seen):
                match_right[right] = left
                return True
        return False

    for left in range(len(adjacency)):
        augment(left, [False] * right_size)
    return match_right


def classify_unmatched_rows(
    only_r: Sequence[Row],
    only_py: Sequence[Row],
    columns: Sequence[str],
    location: str,
    report: ComparisonReport,
) -> None:
    """Pair up the rows each side has exclusively and record them as accepted or rejected.

    Rows are matched 1:1 by maximum bipartite matching over "these two rows differ only in
    accepted-shape ways". Anything left unmatched on either side is a real difference.

    Args:
        only_r: Rows present in R's output but not Python's.
        only_py: Rows present in Python's output but not R's.
        columns: Column names, for readable field descriptions.
        location: The artifact these rows belong to.
        report: The report to append findings to.
    """
    adjacency = [
        [
            j
            for j, py_row in enumerate(only_py)
            if describe_row_difference(r_row, py_row, columns)[0]
        ]
        for r_row in only_r
    ]
    match_right = _maximum_matching(adjacency, len(only_py))

    matched_left = set()
    for right, left in enumerate(match_right):
        if left == -1:
            continue
        matched_left.add(left)
        _, descriptions = describe_row_difference(only_r[left], only_py[right], columns)
        report.accepted.append(RowDivergence(location, descriptions))

    for index, r_row in enumerate(only_r):
        if index not in matched_left:
            report.rejected.append(
                RowDivergence(f"{location} [only in R]", [_render_row(r_row, columns)])
            )
    for index, py_row in enumerate(only_py):
        if match_right[index] == -1:
            report.rejected.append(
                RowDivergence(f"{location} [only in Python]", [_render_row(py_row, columns)])
            )


def classify_set_rows(
    only_r: Sequence[Row],
    only_py: Sequence[Row],
    r_rows: Sequence[Row],
    py_rows: Sequence[Row],
    columns: Sequence[str],
    location: str,
    report: ComparisonReport,
) -> None:
    """Classify differing rows of a *set*-valued artifact (a unique-value list).

    Unlike the processed TSVs — where every row is a record and must pair 1:1 — a unique-list
    sheet is a set, so an accepted divergence can **collapse** two R entries into one Python
    entry (R keeps ``philippines`` and ``philippines r`` apart; the policy makes them the same
    value). The sheet then legitimately has fewer rows. A row is therefore accepted when it
    differs by the accepted shape from *any* row on the other side, matched or not.

    Args:
        only_r: Rows present only in R's sheet.
        only_py: Rows present only in Python's sheet.
        r_rows: Every row of R's sheet.
        py_rows: Every row of Python's sheet.
        columns: Column names, for readable field descriptions.
        location: The artifact these rows belong to.
        report: The report to append findings to.
    """
    for r_row in only_r:
        match = _closest_accepted(r_row, py_rows, columns, r_is_left=True)
        if match is None:
            report.rejected.append(
                RowDivergence(f"{location} [only in R]", [_render_row(r_row, columns)])
            )
        else:
            _, descriptions = describe_row_difference(r_row, match, columns)
            report.accepted.append(RowDivergence(location, descriptions))
    for py_row in only_py:
        match = _closest_accepted(py_row, r_rows, columns, r_is_left=False)
        if match is None:
            report.rejected.append(
                RowDivergence(f"{location} [only in Python]", [_render_row(py_row, columns)])
            )
        else:
            _, descriptions = describe_row_difference(match, py_row, columns)
            report.accepted.append(RowDivergence(location, descriptions))


def _closest_accepted(
    row: Row, candidates: Sequence[Row], columns: Sequence[str], *, r_is_left: bool
) -> Row | None:
    """Find the accepted-shape counterpart of ``row`` that differs from it least.

    The subsequence relation is loose enough to admit several candidates (``asia`` is a
    subsequence of ``australia r`` just as ``australia`` is), so taking the first match would
    report a misleading pairing. Choosing the candidate with the fewest deleted characters —
    ties broken lexicographically — keeps the report both truthful and deterministic.

    Args:
        row: The row to explain.
        candidates: Rows on the other side to match against.
        columns: Column names, for the accepted-shape test.
        r_is_left: Whether ``row`` is the R row (``candidates`` are then Python rows).

    Returns:
        The closest accepted-shape counterpart, or ``None`` if there is none.
    """
    best: Row | None = None
    best_key: tuple[int, Row] | None = None
    for candidate in candidates:
        if candidate == row:
            continue
        pair = (row, candidate) if r_is_left else (candidate, row)
        if not describe_row_difference(pair[0], pair[1], columns)[0]:
            continue
        deleted = sum(len(a) - len(b) for a, b in zip(pair[0], pair[1], strict=True))
        key = (deleted, candidate)
        if best_key is None or key < best_key:
            best, best_key = candidate, key
    return best


def _render_row(row: Row, columns: Sequence[str]) -> str:
    """Render a row as ``col=value`` pairs, skipping empty cells.

    Args:
        row: The row to render.
        columns: Column names.

    Returns:
        A compact single-line rendering.
    """
    parts = []
    for index, cell in enumerate(row):
        if cell == "":
            continue
        name = columns[index] if index < len(columns) else f"col{index}"
        parts.append(f"{name}={cell!r}")
    return " ".join(parts)


# ---------------------------------------------------------------------------
# Preconditions
# ---------------------------------------------------------------------------


def resolve_r_project(explicit: Path | None = None) -> Path:
    """Locate the sibling R project.

    Args:
        explicit: An explicit path, overriding discovery.

    Returns:
        The R project root.

    Raises:
        PreconditionError: If the project is absent or is missing its entry point.
    """
    env = os.environ.get("WHEP_R_PROJECT")
    root = explicit or (Path(env) if env else _DEFAULT_R_PROJECT)
    root = root.expanduser().resolve()
    if not root.is_dir():
        raise PreconditionError(
            f"R project not found at {root}.\n"
            "R is the parity oracle: the reference output cannot be produced without it.\n"
            "Clone the sibling repo whep-digitalization next to this one, or point "
            "--r-project / WHEP_R_PROJECT at it."
        )
    if not (root / _R_ENTRY_POINT).is_file():
        raise PreconditionError(
            f"{root} is not the whep-digitalization project: {_R_ENTRY_POINT} is missing."
        )
    return root


def resolve_rscript() -> Path:
    """Locate the ``Rscript`` executable.

    Returns:
        The path to ``Rscript``.

    Raises:
        PreconditionError: If no R installation can be found.
    """
    env_home = os.environ.get("WHEP_R_HOME")
    home = Path(env_home) if env_home else _DEFAULT_R_HOME
    for name in ("Rscript.exe", "Rscript"):
        exe = home / "bin" / name
        if exe.is_file():
            return exe
    on_path = shutil.which("Rscript")
    if on_path:
        return Path(on_path)
    raise PreconditionError(
        f"Rscript not found (looked in {home / 'bin'} and on PATH).\n"
        "R is the parity oracle and must be executable to produce the reference output.\n"
        "Install R 4.6.0 or point WHEP_R_HOME at an existing installation."
    )


def resolve_dataset(r_project: Path) -> Path:
    """Locate the full raw dataset inside the R project.

    Args:
        r_project: The R project root.

    Returns:
        The ``data/1-import`` directory holding the raw workbooks.

    Raises:
        PreconditionError: If the raw layer is absent or empty.
    """
    import_dir = r_project / "data" / "1-import"
    raw_dir = import_dir / _R_RAW_LAYER
    if not raw_dir.is_dir():
        raise PreconditionError(
            f"Full dataset not found: {raw_dir} does not exist.\n"
            "The full-dataset parity check needs the production workbooks, not the "
            "6-workbook fixture corpus."
        )
    if not any(raw_dir.rglob("*.xlsx")):
        raise PreconditionError(f"Full dataset at {raw_dir} contains no .xlsx workbooks.")
    return import_dir


def count_workbooks(import_dir: Path) -> int:
    """Count the raw workbooks in an R-layout import directory.

    Args:
        import_dir: The ``data/1-import`` directory.

    Returns:
        The number of ``.xlsx`` files under the raw layer.
    """
    return sum(1 for _ in (import_dir / _R_RAW_LAYER).rglob("*.xlsx"))


def newest_mtime(paths: Iterable[Path]) -> float:
    """Return the newest modification time across ``paths``.

    Args:
        paths: Files to inspect.

    Returns:
        The maximum mtime, or ``0.0`` when ``paths`` is empty.
    """
    return max((p.stat().st_mtime for p in paths), default=0.0)


# ---------------------------------------------------------------------------
# Running the two pipelines
# ---------------------------------------------------------------------------


def run_r_pipeline(r_project: Path) -> None:
    """Execute the R project's documented entry point to regenerate the reference output.

    Args:
        r_project: The R project root.

    Raises:
        PreconditionError: If the R run fails.
    """
    rscript = resolve_rscript()
    print(f"[r] {rscript} {_R_ENTRY_POINT}  (cwd={r_project})", flush=True)
    completed = subprocess.run(
        [str(rscript), _R_ENTRY_POINT],
        cwd=r_project,
        check=False,
    )
    if completed.returncode != 0:
        raise PreconditionError(
            f"The R pipeline exited {completed.returncode}; no usable reference output. "
            "Fix the R run before re-running the parity harness."
        )


def stage_python_inputs(import_dir: Path, dst_root: Path) -> None:
    """Copy the R-layout dataset into a throwaway Python project root.

    Args:
        import_dir: The R ``data/1-import`` directory.
        dst_root: The temporary project root to populate.
    """
    dst_import = dst_root / "data" / "import"
    dst_import.mkdir(parents=True, exist_ok=True)
    for r_name, layer in _R_LAYER_MAP.items():
        source = import_dir / r_name
        if source.is_dir():
            shutil.copytree(source, dst_import / layer, dirs_exist_ok=True)


def run_python_pipeline(root: Path) -> None:
    """Run the Python pipeline over a staged project root.

    ``whep_digitize`` is imported here rather than at module scope so the harness runs straight
    from a checkout (``src`` is put on ``sys.path``, as pytest's ``pythonpath`` does) without an
    editable install, and so the precondition checks and the divergence classifier stay
    importable without pulling in the whole pipeline.

    Args:
        root: The staged project root.
    """
    src = _REPO_ROOT / "src"
    if src.is_dir() and str(src) not in sys.path:
        sys.path.insert(0, str(src))

    from whep_digitize.pipeline import run_pipeline
    from whep_digitize.setup.options import RuntimeOptions

    print(f"[py] run_pipeline(root={root})", flush=True)
    run_pipeline(root=root, show_view=False, options=RuntimeOptions(progress_enabled=False))


# ---------------------------------------------------------------------------
# Processed TSV comparison
# ---------------------------------------------------------------------------


def read_tsv(path: Path) -> tuple[str, list[str], list[Row]]:
    """Read a TSV into its line terminator, header, and data rows.

    Args:
        path: The TSV to read.

    Returns:
        ``(terminator, header, rows)``.
    """
    text = path.read_bytes().decode("utf-8")
    terminator = "\r\n" if "\r\n" in text else "\n"
    lines = text.split("\n")
    if lines and lines[-1] == "":
        lines.pop()
    stripped = [line[:-1] if line.endswith("\r") else line for line in lines]
    header = stripped[0].split("\t") if stripped else []
    rows = [tuple(line.split("\t")) for line in stripped[1:]]
    return terminator, header, rows


def value_sum(rows: Iterable[Row], index: int) -> Decimal:
    """Sum a numeric column exactly, ignoring unparseable cells.

    Args:
        rows: The data rows.
        index: The column index to sum.

    Returns:
        The total as an exact :class:`~decimal.Decimal`.
    """
    total = Decimal(0)
    for row in rows:
        if index < len(row):
            try:
                total += Decimal(row[index])
            except (InvalidOperation, ValueError):
                continue
    return total


def compare_processed_tsv(
    r_path: Path, py_path: Path, report: ComparisonReport, budget: int
) -> None:
    """Diff one processed TSV byte-for-byte, classifying any row-level differences.

    Args:
        r_path: The R reference TSV.
        py_path: The Python TSV.
        report: The report to append findings to.
        budget: Accepted differing-row budget, used to cap classification work.
    """
    name = r_path.name
    if r_path.read_bytes() == py_path.read_bytes():
        report.identical.append(name)
        return

    r_terminator, r_header, r_rows = read_tsv(r_path)
    py_terminator, py_header, py_rows = read_tsv(py_path)

    if r_terminator != py_terminator:
        report.errors.append(
            f"{name}: line terminator differs — R {r_terminator!r} vs Python {py_terminator!r}"
        )
    if r_header != py_header:
        report.errors.append(f"{name}: header differs — R {r_header} vs Python {py_header}")
        return
    if len(r_rows) != len(py_rows):
        report.errors.append(
            f"{name}: row count differs — R {len(r_rows)} vs Python {len(py_rows)}"
        )
        return

    if "value" in r_header:
        index = r_header.index("value")
        r_total, py_total = value_sum(r_rows, index), value_sum(py_rows, index)
        if r_total != py_total:
            report.errors.append(
                f"{name}: value sum differs — R {r_total} vs Python {py_total} "
                f"(delta {py_total - r_total})"
            )

    r_counts, py_counts = Counter(r_rows), Counter(py_rows)
    only_r = list((r_counts - py_counts).elements())
    only_py = list((py_counts - r_counts).elements())
    if not only_r and not only_py:
        report.errors.append(
            f"{name}: rows are identical as a multiset but the files differ — "
            "row order diverged (R and Python sort differently)"
        )
        return

    limit = max(budget, 1) * _CLASSIFY_LIMIT_FACTOR
    if len(only_r) > limit or len(only_py) > limit:
        report.errors.append(
            f"{name}: {len(only_r)} R-only and {len(only_py)} Python-only rows — far beyond the "
            f"budget of {budget}; not classified"
        )
        return

    print(
        f"[--] {name}: {len(r_rows)} rows, {len(only_r)} R-only / {len(only_py)} Python-only",
        flush=True,
    )
    classify_unmatched_rows(only_r, only_py, r_header, name, report)


def compare_processed_dirs(r_dir: Path, py_dir: Path, budget: int) -> ComparisonReport:
    """Compare every processed TSV in the two export directories.

    Args:
        r_dir: R's ``processed_data`` directory.
        py_dir: Python's ``processed`` directory.
        budget: Accepted differing-row budget.

    Returns:
        The comparison report.
    """
    report = ComparisonReport("processed TSVs (byte-level)")
    r_files = {p.name: p for p in sorted(r_dir.glob("*.tsv"))}
    py_files = {p.name: p for p in sorted(py_dir.glob("*.tsv"))}
    for missing in sorted(r_files.keys() - py_files.keys()):
        report.errors.append(f"{missing}: present in R output, missing from Python output")
    for extra in sorted(py_files.keys() - r_files.keys()):
        report.errors.append(f"{extra}: present in Python output, missing from R output")
    for name in sorted(r_files.keys() & py_files.keys()):
        compare_processed_tsv(r_files[name], py_files[name], report, budget)
    return report


# ---------------------------------------------------------------------------
# Unique-list workbook comparison (content level)
# ---------------------------------------------------------------------------


def _cell_text(value: Any) -> str:
    """Render a cell value as text for comparison.

    Args:
        value: The raw cell value.

    Returns:
        The value as a string; ``None`` becomes the empty string.
    """
    return "" if value is None else str(value)


def read_workbook_content(path: Path) -> dict[str, list[Row]]:
    """Read a workbook's logical content: ordered sheets, each an ordered list of row tuples.

    Args:
        path: The workbook to read.

    Returns:
        Mapping of sheet name to its rows, cells rendered as text.
    """
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        return {
            sheet: [
                tuple(_cell_text(cell) for cell in row)
                for row in workbook[sheet].iter_rows(values_only=True)
            ]
            for sheet in workbook.sheetnames
        }
    finally:
        workbook.close()


def compare_workbook(r_path: Path, py_path: Path, report: ComparisonReport, budget: int) -> None:
    """Diff one unique-list workbook by sheet names and cell values.

    Args:
        r_path: The R reference workbook.
        py_path: The Python workbook.
        report: The report to append findings to.
        budget: Accepted differing-row budget, used to cap classification work.
    """
    name = r_path.name
    r_content = read_workbook_content(r_path)
    py_content = read_workbook_content(py_path)

    if list(r_content) != list(py_content):
        report.errors.append(
            f"{name}: sheet names differ — R {list(r_content)} vs Python {list(py_content)}"
        )
        return

    clean = True
    for sheet, r_rows in r_content.items():
        py_rows = py_content[sheet]
        location = f"{name}!{sheet}"
        if r_rows == py_rows:
            continue
        clean = False
        # The unique-list sheets carry no header row (values start at row 1), so name the single
        # column after the workbook: unique_polity.xlsx -> "polity".
        columns = [r_path.stem.removeprefix("unique_")]
        r_counts, py_counts = Counter(r_rows), Counter(py_rows)
        only_r = list((r_counts - py_counts).elements())
        only_py = list((py_counts - r_counts).elements())
        if not only_r and not only_py:
            report.errors.append(
                f"{location}: rows are identical as a multiset but ordered differently"
            )
            continue
        limit = max(budget, 1) * _CLASSIFY_LIMIT_FACTOR
        if len(only_r) > limit or len(only_py) > limit:
            report.errors.append(
                f"{location}: {len(only_r)} R-only and {len(only_py)} Python-only rows — "
                f"far beyond the budget of {budget}; not classified"
            )
            continue
        # A row-count difference is expected here and is NOT itself an error: collapsing two R
        # entries onto one Python entry is what the accepted divergence does to a set. It is
        # caught anyway if the collapsed rows do not have the accepted shape.
        print(
            f"[--] {location}: R {len(r_rows)} rows / Python {len(py_rows)} rows, "
            f"{len(only_r)} R-only / {len(only_py)} Python-only",
            flush=True,
        )
        classify_set_rows(only_r, only_py, r_rows, py_rows, columns, location, report)

    if clean:
        report.identical.append(name)


def compare_lists_dirs(r_dir: Path, py_dir: Path, budget: int) -> ComparisonReport:
    """Compare every unique-list workbook in the two export directories.

    Args:
        r_dir: R's ``lists`` directory.
        py_dir: Python's ``lists`` directory.
        budget: Accepted differing-row budget.

    Returns:
        The comparison report.
    """
    report = ComparisonReport("unique-list workbooks (content-level)")
    r_files = {p.name: p for p in sorted(r_dir.glob("*.xlsx"))}
    py_files = {p.name: p for p in sorted(py_dir.glob("*.xlsx"))}
    for missing in sorted(r_files.keys() - py_files.keys()):
        report.errors.append(f"{missing}: present in R output, missing from Python output")
    for extra in sorted(py_files.keys() - r_files.keys()):
        report.errors.append(f"{extra}: present in Python output, missing from R output")
    for name in sorted(r_files.keys() & py_files.keys()):
        compare_workbook(r_files[name], py_files[name], report, budget)
    return report


# ---------------------------------------------------------------------------
# Orchestration
# ---------------------------------------------------------------------------


@dataclass
class ParityOutcome:
    """The full-dataset parity result.

    Attributes:
        workbooks: Number of raw workbooks compared.
        reports: One report per artifact class.
        budget: The accepted differing-row budget.
    """

    workbooks: int
    reports: list[ComparisonReport]
    budget: int

    @property
    def accepted_rows(self) -> int:
        """Total rows differing only by the accepted normalization policy."""
        return sum(len(r.accepted) for r in self.reports)

    @property
    def failures(self) -> list[str]:
        """Every reason the run must exit non-zero."""
        reasons: list[str] = []
        for report in self.reports:
            reasons.extend(f"[{report.label}] {e}" for e in report.errors)
            reasons.extend(f"[{report.label}] {d.render()}" for d in report.rejected)
            if len(report.accepted) > self.budget:
                reasons.append(
                    f"[{report.label}] {len(report.accepted)} accepted-shape differing rows "
                    f"exceed the budget of {self.budget}"
                )
        return reasons


def check_parity(
    r_project: Path,
    *,
    run_r: bool,
    budget: int,
    allow_stale: bool,
    keep_output: bool,
) -> ParityOutcome:
    """Run both pipelines over the full dataset and compare their outputs.

    Args:
        r_project: The R project root.
        run_r: Regenerate the R reference by executing the R pipeline.
        budget: Accepted differing-row budget.
        allow_stale: Accept an R reference older than the newest input.
        keep_output: Leave the staged Python output on disk.

    Returns:
        The parity outcome.

    Raises:
        PreconditionError: If the R reference is absent or stale.
    """
    import_dir = resolve_dataset(r_project)
    workbooks = count_workbooks(import_dir)
    print(f"[--] full dataset: {workbooks} raw workbooks under {import_dir / _R_RAW_LAYER}")

    if run_r:
        run_r_pipeline(r_project)

    r_processed = r_project / _R_PROCESSED_DIR
    r_lists = r_project / _R_LISTS_DIR
    if not r_processed.is_dir() or not any(r_processed.glob("*.tsv")):
        raise PreconditionError(
            f"No R reference output at {r_processed}.\n"
            "R is the parity oracle — re-run with --run-r to generate it."
        )

    reference_mtime = min(p.stat().st_mtime for p in r_processed.glob("*.tsv"))
    input_mtime = newest_mtime(import_dir.rglob("*.xlsx"))
    if reference_mtime < input_mtime and not allow_stale:
        raise PreconditionError(
            f"The R reference output at {r_processed} predates the newest input workbook; "
            "it does not describe the current dataset.\n"
            "Re-run with --run-r to regenerate it, or --allow-stale-r-output to compare anyway."
        )

    tmp_root = Path(tempfile.mkdtemp(prefix="whep_parity_"))
    try:
        stage_python_inputs(import_dir, tmp_root)
        run_python_pipeline(tmp_root)
        reports = [
            compare_processed_dirs(r_processed, tmp_root / _PY_PROCESSED_DIR, budget),
            compare_lists_dirs(r_lists, tmp_root / _PY_LISTS_DIR, budget),
        ]
    finally:
        if keep_output:
            print(f"[--] staged Python output kept at {tmp_root}")
        else:
            shutil.rmtree(tmp_root, ignore_errors=True)

    return ParityOutcome(workbooks=workbooks, reports=reports, budget=budget)


def print_outcome(outcome: ParityOutcome) -> None:
    """Print a human-readable parity summary.

    Args:
        outcome: The parity outcome.
    """
    print(f"\n{'=' * 78}\nFull-dataset parity — {outcome.workbooks} workbooks\n{'=' * 78}")
    for report in outcome.reports:
        print(f"\n{report.label}")
        print(f"  identical files      : {len(report.identical)} {report.identical}")
        print(f"  accepted diff rows   : {len(report.accepted)} (budget {outcome.budget})")
        print(f"  rejected diff rows   : {len(report.rejected)}")
        print(f"  structural errors    : {len(report.errors)}")
        for divergence in report.accepted[:20]:
            print(f"    ~ {divergence.render()}")
        if len(report.accepted) > 20:
            print(f"    ~ ... and {len(report.accepted) - 20} more")

    failures = outcome.failures
    if failures:
        print(f"\nFAIL — {len(failures)} problem(s):")
        for reason in failures[:50]:
            print(f"  ! {reason}")
        if len(failures) > 50:
            print(f"  ! ... and {len(failures) - 50} more")
    else:
        print(
            f"\nPASS — outputs match R except {outcome.accepted_rows} accepted "
            "normalization-policy row(s); row counts and value sums identical."
        )


def main(argv: Sequence[str] | None = None) -> int:
    """Entry point.

    Args:
        argv: Command-line arguments; defaults to ``sys.argv[1:]``.

    Returns:
        ``0`` on parity, ``1`` on a parity failure, ``2`` on a precondition failure.
    """
    parser = argparse.ArgumentParser(
        description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter
    )
    parser.add_argument("--r-project", type=Path, default=None, help="R project root")
    parser.add_argument(
        "--run-r",
        action="store_true",
        help="regenerate the R reference (OVERWRITES the R project's "
        "data/2-postpro and data/3-export)",
    )
    parser.add_argument(
        "--divergence-budget",
        type=int,
        default=_DEFAULT_DIVERGENCE_BUDGET,
        help=f"accepted differing-row budget (default {_DEFAULT_DIVERGENCE_BUDGET})",
    )
    parser.add_argument(
        "--allow-stale-r-output",
        action="store_true",
        help="compare against an R reference older than the newest input workbook",
    )
    parser.add_argument("--keep-output", action="store_true", help="keep the staged Python output")
    args = parser.parse_args(argv)

    try:
        r_project = resolve_r_project(args.r_project)
        if args.run_r:
            resolve_rscript()
        outcome = check_parity(
            r_project,
            run_r=args.run_r,
            budget=args.divergence_budget,
            allow_stale=args.allow_stale_r_output,
            keep_output=args.keep_output,
        )
    except PreconditionError as exc:
        print(f"\nPRECONDITION FAILED\n{exc}", file=sys.stderr)
        return 2

    print_outcome(outcome)
    return 1 if outcome.failures else 0


if __name__ == "__main__":
    raise SystemExit(main())
